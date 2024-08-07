import os
import json
from dotenv import load_dotenv
from linkedin_api import Linkedin
from openpyxl import load_workbook
from pymongo import MongoClient
import pandas as pd
import streamlit as st
from requests.cookies import cookiejar_from_dict

load_dotenv()
#3.10
# Set LinkedIn cookies directly
os.environ["LINKEDIN_COOKIE_LI_AT"] = "AQEDASar904EhzN7AAABkJEpS-wAAAGQtTXP7E0AUMqyNC4s97XRXHZChzLJjFbbcZjmzVUpyMiuPI-9RxUjMxRXQuPKE-3shYFtz6E_F-MFCGwUr_ZrslJYajHWQl0QoMG1NiAUfDdXae-G6snwvmEF"
os.environ["LINKEDIN_COOKIE_JSESSIONID"] = "ajax:7756477511703306536"
username = "nileshk191@gmail.com"
password = "Nilu@191"
# Function to append data to Excel
def append_to_excel(profile_data, sheet_name):
    wb = load_workbook('https://github.com/Nilesh-inProspects/web-Scraping/blob/main/faculty3.xlsx')
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        # Create headers based on the sheet name
        if sheet_name == "LinkedIn Data 1":
            ws.append(["Full Name", "Profile Headline", "Industry", "Location", "Summary", "Skills", "Profile URL"])
        elif sheet_name == "LinkedIn Data 2":
            ws.append(["Full Name", "School", "Degree", "Field of study", "Start dateMonth", "Start dateYear",
                       "End dateMonth", "End dateYear", "Grade", "Activities and societies", "Description", "Skills", "Media",
                       "Type", "Location", "Start dateMonth", "Start dateYear", "End dateMonth", "End dateYear",
                       "Description", "Profile Headline", "Media(Links)", "Media(pictures)"])
        elif sheet_name == "LinkedIn Data 3":
            ws.append(["Full Name", "License or certification", "Project", "Course"])
        elif sheet_name == "LinkedIn Data 4":
            ws.append(["Full Name", "Volunteer experience", "Publications", "Patent"])

    full_name = f"{profile_data['firstName']} {profile_data['lastName']}"
    profile_headline = profile_data.get('headline', '')
    industry = profile_data.get('industryName', '')
    location = profile_data.get('locationName', '')
    summary = profile_data.get('summary', '')
    skills = ', '.join([skill['name'] for skill in profile_data.get('skills', [])])
    profile_url = f"https://www.linkedin.com/in/{profile_data['public_id']}"

    if sheet_name == "LinkedIn Data 1":
        ws.append([full_name, profile_headline, industry, location, summary, skills, profile_url])

    if sheet_name == "LinkedIn Data 2":
        education = profile_data.get('education', [])
        for edu in education:
            ws.append([
                full_name,
                edu.get('school', {}).get('schoolName', ''),
                edu.get('degreeName', ''),
                edu.get('fieldOfStudy', ''),
                edu.get('timePeriod', {}).get('startDate', {}).get('month', ''),
                edu.get('timePeriod', {}).get('startDate', {}).get('year', ''),
                edu.get('timePeriod', {}).get('endDate', {}).get('month', ''),
                edu.get('timePeriod', {}).get('endDate', {}).get('year', ''),
                edu.get('grade', ''),
                '',  # Activities and societies not provided
                edu.get('description', ''),
                skills,
                '',  # Media not provided
                '',  # Type not provided
                location,
                '',  # Start date month not provided
                '',  # Start date year not provided
                '',  # End date month not provided
                '',  # End date year not provided
                '',  # Description not provided
                profile_headline,
                '',  # Media links not provided
                ''   # Media pictures not provided
            ])

    if sheet_name == "LinkedIn Data 3":
        certifications = profile_data.get('certifications', [])
        for cert in certifications:
            ws.append([
                full_name,
                cert.get('name', ''),
                cert.get('authority', ''),
                cert.get('timePeriod', {}).get('startDate', {}).get('month', ''),
                cert.get('timePeriod', {}).get('startDate', {}).get('year', ''),
                cert.get('timePeriod', {}).get('endDate', {}).get('month', ''),
                cert.get('timePeriod', {}).get('endDate', {}).get('year', ''),
                cert.get('licenseNumber', ''),
                cert.get('url', ''),
                skills,
                '',  # Media not provided
                '',  # Project name not provided
                '',  # Project description not provided
                '',  # Project skills not provided
                '',  # Project media not provided
                '',  # Currently working on project not provided
                '',  # Start date month not provided
                '',  # Start date year not provided
                '',  # End date month not provided
                '',  # End date year not provided
                '',  # Add contributors not provided
                '',  # Associated with not provided
                '',  # Course name not provided
                '',  # Number not provided
                ''   # Associated with not provided
            ])

    if sheet_name == "LinkedIn Data 4":
        volunteer = profile_data.get('volunteer', [])
        publications = profile_data.get('publications', [])
        patents = profile_data.get('patents', [])
        for vol in volunteer:
            ws.append([
                full_name,
                vol.get('organization', ''),
                vol.get('role', ''),
                vol.get('cause', ''),
                vol.get('timePeriod', {}).get('startDate', {}).get('month', ''),
                vol.get('timePeriod', {}).get('startDate', {}).get('year', ''),
                vol.get('timePeriod', {}).get('endDate', {}).get('month', ''),
                vol.get('timePeriod', {}).get('endDate', {}).get('year', ''),
                vol.get('description', ''),
                '',  # Media not provided
                '',  # Title not provided
                '',  # Publication/Publisher not provided
                '',  # Publication date not provided
                '',  # Author not provided
                '',  # Publication URL not provided
                '',  # Publication description not provided
                '',  # Patent title not provided
                '',  # Patent or application number not provided
                '',  # Inventor not provided
                '',  # Status not provided
                '',  # Issue date not provided
                '',  # Patent URL not provided
                ''   # Patent description not provided
            ])

    wb.save('https://github.com/Nilesh-inProspects/web-Scraping/blob/main/faculty3.xlsx')

# Function to insert data into MongoDB
def insert_to_mongodb(profile_data):
    client = MongoClient('mongodb+srv://nilu191:nilu191@cluster0.slak3.mongodb.net/')
    db = client['linkedin']
    collection = db['profiles']
    collection.insert_one(profile_data)

# Function to process profile data
def process_profile(profile_data):
    append_to_excel(profile_data, "LinkedIn Data 1")
    append_to_excel(profile_data, "LinkedIn Data 2")
    append_to_excel(profile_data, "LinkedIn Data 3")
    append_to_excel(profile_data, "LinkedIn Data 4")
    insert_to_mongodb(profile_data)

# Function to get LinkedIn profile data
def get_linkedin_profile(api, public_id):
    profile = api.get_profile(public_id)
    return profile

# Function to process multiple profiles from an Excel file
def process_multiple_profiles(file):
    df = pd.read_excel(file)
    linkedin_urls = df['Linkedln URL'].tolist()
    cookies = cookiejar_from_dict(
        {
            "liap": "true",
            "li_at": os.environ["LINKEDIN_COOKIE_LI_AT"],
            "JSESSIONID": os.environ["LINKEDIN_COOKIE_JSESSIONID"],
        }
    )
    api = Linkedin(username, password, cookies=cookies)

    for linkedin_url in linkedin_urls:
        public_id = linkedin_url.split('/')[-1]
        profile_data = get_linkedin_profile(api, public_id)
        process_profile(profile_data)

# Main function
def main():
    st.title("LinkedIn Profile Data Fetcher")
    uploaded_file = st.file_uploader("Choose an Excel file with LinkedIn URLs", type="xlsx")
    if uploaded_file is not None:
        st.write("Processing...")
        process_multiple_profiles(uploaded_file)
        st.write("Data successfully appended to Excel and MongoDB!")

if __name__ == "__main__":
    main()
